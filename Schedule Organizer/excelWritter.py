import openpyxl
# import Employee

'''
rows:
2 = header1
20 = header2 SCH
23 = header3 HC
27 = header4 SC
30 = header5 BAG
33 = header6 CART
'''
# 0             1              2       3        4          5            6
# data_Cajeros, data_Training, data_SC, data_HC, data_SCH, data_Baggers, data_Cart
def write_Plan(data: list(list())):

    wb = openpyxl.load_workbook('./Plan de Accion Base.xlsx')

    sheet = wb.active

    date = 'Fecha:' # add week date

    sheet.cell(row=1, column=1, value=date)

    r = 3
    for emps in data[0]:
        sheet.cell(row=r, column=1, value=emps.name)
        sheet.cell(row=r, column=2, value=emps.str_entry)
        sheet.cell(row=r, column=3, value=emps.str_out)
        r += 1

    if data[1]:
        training_Header = len(data[0]) + 4
        r = training_Header + 1
        sheet.cell(row=training_Header, column=1, value='Training')

        for emps in data[1]:
            sheet.cell(row=r, column=1, value=emps.name)
            sheet.cell(row=r, column=2, value=emps.str_entry)
            sheet.cell(row=r, column=3, value=emps.str_out)
            r += 1

    r = 21
    for emps in data[4]:
            sheet.cell(row=r, column=1, value=emps.name)
            sheet.cell(row=r, column=2, value=emps.str_entry)
            sheet.cell(row=r, column=3, value=emps.str_out)
            r += 1
    
    r = 24
    for emps in data[3]:
        sheet.cell(row=r, column=1, value=emps.name)
        sheet.cell(row=r, column=2, value=emps.str_entry)
        sheet.cell(row=r, column=3, value=emps.str_out)
        r += 1
    
    r = 28
    for emps in data[2]:
        sheet.cell(row=r, column=1, value=emps.name)
        sheet.cell(row=r, column=2, value=emps.str_entry)
        sheet.cell(row=r, column=3, value=emps.str_out)
        r += 1
    
    r = 31
    for emps in data[5]:
        sheet.cell(row=r, column=1, value=emps.name)
        sheet.cell(row=r, column=2, value=emps.str_entry)
        sheet.cell(row=r, column=3, value=emps.str_out)
        r += 1

    r = 34 
    for emps in data[6]:
        sheet.cell(row=r, column=1, value=emps.name)
        sheet.cell(row=r, column=2, value=emps.str_entry)
        sheet.cell(row=r, column=3, value=emps.str_out)
        r += 1
    
    wb.save('Plan2.xlsx')





